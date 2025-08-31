import pandas as pd
from openpyxl import load_workbook
import sys

def analisar_excel(caminho_arquivo, aba_nome='Estoque'):
    """Analisa a estrutura de um arquivo Excel"""
    try:
        print(f"📊 Analisando arquivo: {caminho_arquivo}")
        print(f"📋 Aba: {aba_nome}")
        
        # Verificar abas disponíveis
        wb = load_workbook(caminho_arquivo, read_only=True)
        print(f"📑 Abas disponíveis: {', '.join(wb.sheetnames)}")
        
        # Ler dados
        df = pd.read_excel(caminho_arquivo, sheet_name=aba_nome)
        print(f"✅ Arquivo lido com sucesso")
        print(f"📈 Total de linhas: {len(df)}")
        print(f"📊 Total de colunas: {len(df.columns)}")
        
        print("\n🔍 Colunas encontradas:")
        for i, coluna in enumerate(df.columns, 1):
            print(f"  {i}. {coluna} (Tipo: {type(coluna)})")
        
        print("\n📝 Primeiras linhas de dados:")
        print(df.head().to_string())
        
        print("\n💡 Dicas:")
        print("- Verifique se há colunas como 'Nome', 'Número do Bem', 'Localização'")
        print("- Os nomes das colunas são sensíveis a acentos e maiúsculas/minúsculas")
        print("- Se necessário, renomeie as colunas no Excel para os nomes padrão")
        
    except Exception as e:
        print(f"❌ Erro ao analisar arquivo: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        arquivo = sys.argv[1]
        aba = sys.argv[2] if len(sys.argv) > 2 else 'Estoque'
        analisar_excel(arquivo, aba)
    else:
        print("Uso: python debug_excel.py caminho/do/arquivo.xlsx [nome_da_aba]")