import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Monta o caminho de forma confiável
CAMINHO_PLANILHA = os.path.join('relatorios', 'controle_patrimonial.xlsx')

def verificar_patrimonio(patrimonio, caminho_planilha=CAMINHO_PLANILHA):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip() == patrimonio.strip():
            return True
    return False

def marcar_localizado(patrimonio, caminho_planilha=CAMINHO_PLANILHA):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip() == patrimonio.strip():
            for cell in row:
                cell.fill = verde
            # Garante que a coluna de setor localizado existe
            if len(row) >= 4:
                row[3].value = "Localizado"
            break
    wb.save(caminho_planilha)
print("Caminho absoluto:", os.path.abspath(CAMINHO_PLANILHA))
print("Arquivo existe?", os.path.exists(CAMINHO_PLANILHA))
