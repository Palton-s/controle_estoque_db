from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ABA_ESTOQUE = "Estoque"

def verificar_bem(numero_bem, caminho_planilha):
    try:
        wb = load_workbook(caminho_planilha, read_only=True, data_only=True)
    except Exception as e:
        print(f"[ERRO ao carregar planilha]: {e}")
        return False, "Erro ao abrir a planilha."

    if ABA_ESTOQUE not in wb.sheetnames:
        return False, f"Aba '{ABA_ESTOQUE}' não encontrada na planilha."

    ws = wb[ABA_ESTOQUE]
    numero_bem_limpo = str(numero_bem).strip().lstrip("0")

    for row in ws.iter_rows(min_row=2):
        valor_celula = row[1].value
        if valor_celula:
            valor_celula_limpo = str(valor_celula).strip().lstrip("0")
            print(f"[DEBUG] Comparando '{valor_celula_limpo}' com '{numero_bem_limpo}'")
            if valor_celula_limpo == numero_bem_limpo:
                return True, None

    return False, f"Bem {numero_bem} não encontrado na aba '{ABA_ESTOQUE}'."

def marcar_bem_localizado(numero_bem, caminho_planilha, localizacao=None):
    try:
        wb = load_workbook(caminho_planilha)
    except Exception as e:
        print(f"[ERRO ao abrir planilha para marcação]: {e}")
        return "Erro ao abrir a planilha para marcação."

    if ABA_ESTOQUE not in wb.sheetnames:
        return f"Aba '{ABA_ESTOQUE}' não encontrada na planilha."

    ws = wb[ABA_ESTOQUE]
    numero_bem_limpo = str(numero_bem).strip().lstrip("0")
    alterado = False

    for row in ws.iter_rows(min_row=2):
        valor_celula = row[1].value
        if valor_celula:
            valor_celula_limpo = str(valor_celula).strip().lstrip("0")
            if valor_celula_limpo == numero_bem_limpo:
                if len(row) >= 3:
                    row[2].value = "OK"
                if localizacao and len(row) >= 4:
                    row[3].value = localizacao
                for cell in row:
                    cell.fill = verde
                alterado = True
                break

    if alterado:
        try:
            wb.save(caminho_planilha)
            return f"✅ Bem {numero_bem} marcado como localizado com sucesso em '{localizacao}'!"
        except Exception as e:
            print(f"[ERRO ao salvar planilha]: {e}")
            return "Erro ao salvar alterações na planilha."
    else:
        return f"Bem {numero_bem} não encontrado na aba '{ABA_ESTOQUE}'."

def buscar_localizacao_existente(numero_bem, caminho_planilha):
    try:
        wb = load_workbook(caminho_planilha, data_only=True)
    except Exception as e:
        print(f"[ERRO ao abrir planilha para buscar localização]: {e}")
        return "localização não disponível"

    if ABA_ESTOQUE not in wb.sheetnames:
        return "aba não encontrada"

    ws = wb[ABA_ESTOQUE]
    numero_bem_limpo = str(numero_bem).strip().lstrip("0")

    for row in ws.iter_rows(min_row=2):
        valor_celula = row[1].value
        if valor_celula:
            valor_celula_limpo = str(valor_celula).strip().lstrip("0")
            if valor_celula_limpo == numero_bem_limpo:
                if len(row) >= 4 and row[3].value:
                    return str(row[3].value).strip()
                else:
                    return "localização não registrada"
    return "bem não encontrado"

def gerar_planilhas_localizacao(caminho_original):
    wb_original = load_workbook(caminho_original, data_only=True)
    if ABA_ESTOQUE not in wb_original.sheetnames:
        return [], []

    ws = wb_original[ABA_ESTOQUE]
    localizados = []
    nao_localizados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        if (
            str(row[0]).strip().lower() == 'nome' and
            str(row[1]).strip().lower() in ['nº do bem', 'numero', 'número'] and
            str(row[2]).strip().lower() == 'situação'
        ):
            continue

        nome = row[0] if len(row) > 0 else ""
        numero = row[1] if len(row) > 1 else ""
        situacao = row[2] if len(row) > 2 else ""

        registro = {
            'nome': nome,
            'numero': numero,
            'situacao': situacao
        }

        if str(situacao).strip().upper() == 'OK':
            localizados.append(registro)
        else:
            nao_localizados.append(registro)

    return localizados, nao_localizados
