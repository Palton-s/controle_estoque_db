import sqlite3
import pandas as pd
from openpyxl import load_workbook
import os
import shutil
from typing import Tuple, List, Dict, Any, Optional
from datetime import datetime, date
import logging

# Configurar logger
logger = logging.getLogger(__name__)

# ==============================
# FUNÇÕES AUXILIARES
# ==============================

def get_db_connection(db_path: str) -> sqlite3.Connection:
    """Conexão simples com o banco SQLite"""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def criar_tabela_atualizada(db_path: str) -> None:
    """Cria a tabela bens com a nova estrutura completa"""
    try:
        conn = get_db_connection(db_path)
        cursor = conn.cursor()
        
        # Criar tabela principal
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT UNIQUE NOT NULL,
                nome TEXT NOT NULL,
                situacao TEXT DEFAULT 'Pendente',
                localizacao TEXT,
                responsavel TEXT,
                data_ultima_vistoria DATE,
                data_vistoria_atual DATE,
                auditor TEXT,
                observacoes TEXT,
                data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                data_localizacao DATETIME
            )
        ''')
        
        # Verificar colunas existentes
        cursor.execute("PRAGMA table_info(bens)")
        colunas_existentes = [col[1] for col in cursor.fetchall()]
        
        # Colunas novas para adicionar se não existirem
        novas_colunas = [
            ('responsavel', 'TEXT'),
            ('data_ultima_vistoria', 'DATE'),
            ('data_vistoria_atual', 'DATE'),
            ('auditor', 'TEXT')
        ]
        
        for coluna, tipo in novas_colunas:
            if coluna not in colunas_existentes:
                try:
                    cursor.execute(f"ALTER TABLE bens ADD COLUMN {coluna} {tipo}")
                    print(f"Coluna {coluna} adicionada à tabela bens")
                except sqlite3.OperationalError:
                    # Coluna já existe, ignorar erro
                    pass
        
        conn.commit()
        conn.close()
        print("Tabela bens criada/atualizada com sucesso")
        
    except Exception as e:
        print(f"Erro ao criar/atualizar tabela: {str(e)}")
        raise

def processar_data_excel(valor: Any) -> date:
    """Processa datas do Excel para formato Python"""
    try:
        if valor is None or pd.isna(valor):
            return datetime.now().date()
        
        if isinstance(valor, datetime):
            return valor.date()
        elif isinstance(valor, date):
            return valor
        elif isinstance(valor, str) and valor.strip():
            # Tentar diferentes formatos de data
            formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d']
            for formato in formatos:
                try:
                    return datetime.strptime(valor.strip(), formato).date()
                except ValueError:
                    continue
            return datetime.now().date()
        else:
            return datetime.now().date()
    except Exception:
        return datetime.now().date()

def normalizar_valor(valor: Any) -> Optional[str]:
    """Normaliza valores para evitar problemas de tipo e formato"""
    if pd.isna(valor) or valor is None:
        return None
    
    try:
        valor_str = str(valor).strip()
        return valor_str if valor_str else None
    except Exception:
        return None

# ==============================
# FUNÇÃO DE DETECÇÃO DE COLUNAS
# ==============================

def detectar_colunas(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Detecta automaticamente as colunas relevantes no DataFrame
    """
    # CORREÇÃO: Remover espaços extras dos nomes das colunas
    df.columns = [str(col).strip() for col in df.columns]
    
    mapeamento_colunas = {
        'numero': None,
        'nome': None,
        'situacao': None,
        'localizacao': None,
        'responsavel': None,
        'data_ultima_vistoria': None,
        'data_vistoria_atual': None,
        'auditor': None
    }
    
    # Mapeamento de possíveis nomes para cada coluna
    possiveis_nomes = {
        'numero': ['número do bem', 'numero do bem', 'nº do bem', 'patrimonio', 'patrimônio', 'numero', 'número'],
        'nome': ['nome', 'descrição', 'descricao', 'item', 'equipamento', 'bem'],
        'situacao': ['situação', 'situacao', 'status', 'estado', 'condição'],
        'localizacao': ['localização', 'localizacao', 'local', 'setor', 'departamento'],
        'responsavel': ['responsavel', 'responsável', 'encarregado', 'curador'],
        'data_ultima_vistoria': ['data da ultima vistoria', 'última vistoria', 'data ultima vistoria'],
        'data_vistoria_atual': ['data da vistoria atual', 'vistoria atual', 'data vistoria atual'],
        'auditor': ['auditor', 'auditor responsavel', 'inspetor']
    }
    
    # Converter nomes das colunas para minúsculas
    colunas_df = [str(col).strip().lower() for col in df.columns]
    
    print(f"🔍 Colunas encontradas no Excel: {df.columns.tolist()}")
    
    # Procurar correspondências
    for coluna_alvo, possibilidades in possiveis_nomes.items():
        for possibilidade in possibilidades:
            for idx, coluna_df in enumerate(colunas_df):
                if possibilidade in coluna_df:
                    mapeamento_colunas[coluna_alvo] = df.columns[idx]
                    print(f"✅ Coluna '{coluna_alvo}' detectada como: '{df.columns[idx]}'")
                    break
            if mapeamento_colunas[coluna_alvo]:
                break
    
    return mapeamento_colunas

# ==============================
# FUNÇÃO PRINCIPAL DE IMPORTAÇÃO
# ==============================

def importar_excel_para_sqlite(caminho_excel: str, aba_nome: str, db_path: str, criar_backup: bool = True) -> Tuple[bool, str]:
    """
    Importa dados do Excel para SQLite com nova estrutura completa
    """
    try:
        print(f"📁 Iniciando importação do arquivo: {caminho_excel}")
        
        # Verificar se arquivo existe
        if not os.path.exists(caminho_excel):
            return False, f"❌ Arquivo Excel não encontrado: {caminho_excel}"
        
        # Criar/atualizar tabela primeiro
        criar_tabela_atualizada(db_path)
        
        # Fazer backup se necessário
        backup_path = ""
        if criar_backup and os.path.exists(db_path):
            backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            backup_path = os.path.join(backup_dir, f"backup_controle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
            shutil.copy2(db_path, backup_path)
            mensagem_backup = f"Backup criado: {os.path.basename(backup_path)}"
            print(f"📦 {mensagem_backup}")
        else:
            mensagem_backup = ""
        
        # Carregar arquivo Excel
        try:
            wb = load_workbook(caminho_excel, read_only=True)
        except Exception as e:
            return False, f"❌ Erro ao abrir arquivo Excel: {str(e)}"
        
        if aba_nome not in wb.sheetnames:
            abas_disponiveis = ", ".join(wb.sheetnames)
            wb.close()
            return False, f"❌ Aba '{aba_nome}' não encontrada. Abas disponíveis: {abas_disponiveis}"
        
        wb.close()
        
        # Ler dados com pandas
        try:
            df = pd.read_excel(caminho_excel, sheet_name=aba_nome)
        except Exception as e:
            return False, f"❌ Erro ao ler dados do Excel: {str(e)}"
        
        if df.empty:
            return False, "❌ O arquivo Excel está vazio ou não contém dados"
        
        print(f"📊 Total de linhas no Excel: {len(df)}")
        
        # Detectar mapeamento de colunas
        mapeamento = detectar_colunas(df)
        
        # Verificar colunas obrigatórias
        if not mapeamento['numero']:
            return False, "❌ Coluna do número do bem não detectada"
        if not mapeamento['nome']:
            return False, "❌ Coluna do nome não detectada"
        
        # Processar e importar dados
        conn = get_db_connection(db_path)
        cursor = conn.cursor()
        
        registros_inseridos = 0
        registros_atualizados = 0
        registros_erro = 0
        
        for index, row in df.iterrows():
            try:
                # Pular linhas vazias
                if row.isnull().all():
                    continue
                
                # Extrair dados básicos
                numero = normalizar_valor(row[mapeamento['numero']])
                nome = normalizar_valor(row[mapeamento['nome']])
                
                if not numero or not nome:
                    registros_erro += 1
                    continue
                
                # Extrair dados opcionais
                situacao_val = row.get(mapeamento.get('situacao'), 'Pendente')
                situacao = normalizar_valor(situacao_val) or 'Pendente'
                
                localizacao_val = row.get(mapeamento.get('localizacao'))
                localizacao = normalizar_valor(localizacao_val) or ''
                
                responsavel_val = row.get(mapeamento.get('responsavel'))
                responsavel = normalizar_valor(responsavel_val) or ''
                
                auditor_val = row.get(mapeamento.get('auditor'))
                auditor = normalizar_valor(auditor_val) or ''
                
                # Processar datas
                data_ultima_val = row.get(mapeamento.get('data_ultima_vistoria'))
                data_ultima_vistoria = processar_data_excel(data_ultima_val)
                
                data_atual_val = row.get(mapeamento.get('data_vistoria_atual'))
                data_vistoria_atual = processar_data_excel(data_atual_val)
                
                # Verificar se registro já existe
                cursor.execute("SELECT id FROM bens WHERE numero = ?", (numero,))
                existe = cursor.fetchone()
                
                if existe:
                    # Atualizar registro existente
                    cursor.execute('''
                        UPDATE bens SET 
                        nome = ?, situacao = ?, localizacao = ?, responsavel = ?,
                        data_ultima_vistoria = ?, data_vistoria_atual = ?, auditor = ?
                        WHERE numero = ?
                    ''', (nome, situacao, localizacao, responsavel, 
                         data_ultima_vistoria, data_vistoria_atual, auditor, numero))
                    registros_atualizados += 1
                else:
                    # Inserir novo registro
                    cursor.execute('''
                        INSERT INTO bens 
                        (numero, nome, situacao, localizacao, responsavel, 
                         data_ultima_vistoria, data_vistoria_atual, auditor)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (numero, nome, situacao, localizacao, responsavel,
                         data_ultima_vistoria, data_vistoria_atual, auditor))
                    registros_inseridos += 1
                
                # Log a cada 50 registros
                if (registros_inseridos + registros_atualizados) % 50 == 0:
                    print(f"📈 Processados: {registros_inseridos + registros_atualizados} registros")
                
            except Exception as e:
                registros_erro += 1
                if registros_erro <= 5:  # Mostrar apenas os primeiros 5 erros
                    print(f"⚠️  Erro na linha {index + 2}: {str(e)}")
                continue
        
        conn.commit()
        conn.close()
        
        # Mensagem de resultado
        mensagem_log = f"✅ Importação concluída! Novos: {registros_inseridos}, Atualizados: {registros_atualizados}"
        if registros_erro > 0:
            mensagem_log += f", Erros: {registros_erro}"
        if mensagem_backup:
            mensagem_log += f" | {mensagem_backup}"
        
        print(mensagem_log)
        
        # Mensagem para usuário
        mensagem_user = f"✅ Importação concluída com sucesso!"
        mensagem_user += f"\n• 📊 Novos registros: {registros_inseridos}"
        mensagem_user += f"\n• 🔄 Registros atualizados: {registros_atualizados}"
        
        if registros_erro > 0:
            mensagem_user += f"\n• ⚠️  Registros com erro: {registros_erro}"
        
        if mensagem_backup:
            mensagem_user += f"\n• 📦 {mensagem_backup}"
        
        return True, mensagem_user
        
    except Exception as e:
        error_msg = f"❌ Erro na importação: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return False, error_msg

# ==============================
# FUNÇÕES DE VERIFICAÇÃO
# ==============================

def verificar_estrutura_excel(caminho_arquivo: str, aba_nome: str = 'Estoque') -> Tuple[bool, str]:
    """Verifica se a planilha tem a estrutura esperada"""
    try:
        if not os.path.exists(caminho_arquivo):
            return False, f"Arquivo não encontrado: {caminho_arquivo}"
        
        wb = load_workbook(caminho_arquivo, read_only=True)
        
        if aba_nome not in wb.sheetnames:
            abas_disponiveis = ", ".join(wb.sheetnames)
            wb.close()
            return False, f"Aba '{aba_nome}' não encontrada. Abas disponíveis: {abas_disponiveis}"
        
        ws = wb[aba_nome]
        
        # Obter cabeçalhos
        cabecalhos = []
        for cell in ws[1]:
            if cell.value:
                cabecalhos.append(str(cell.value).strip().upper())
        
        wb.close()
        
        # Campos obrigatórios
        obrigatorios = ['NOME', 'NUMERO DO BEM', 'SITUAÇÃO', 'LOCALIZAÇÃO']
        
        for campo in obrigatorios:
            if campo not in cabecalhos:
                return False, f"Campo obrigatório '{campo}' não encontrado"
        
        return True, f"Estrutura válida. Campos encontrados: {len(cabecalhos)}"
        
    except Exception as e:
        return False, f"Erro ao verificar estrutura: {str(e)}"

def obter_colunas_excel(arquivo_excel: str, aba_nome: str = 'Estoque') -> List[str]:
    """Retorna as colunas disponíveis no arquivo Excel"""
    try:
        df = pd.read_excel(arquivo_excel, sheet_name=aba_nome, nrows=1)
        return df.columns.tolist()
    except Exception as e:
        print(f"Erro ao obter colunas: {str(e)}")
        return []

# ==============================
# FUNÇÃO PARA TESTE/DEBUG
# ==============================

def testar_importacao():
    """Função para testar a importação diretamente"""
    try:
        print("=== TESTANDO IMPORTAÇÃO EXCEL ===")
        
        # Caminho correto para o arquivo Excel
        excel_path = "relatorios/controle_patrimonial.xlsx"
        db_path = "relatorios/controle_patrimonial.db"
        
        print(f"📁 Procurando arquivo: {excel_path}")
        
        if not os.path.exists(excel_path):
            print(f"❌ Arquivo não encontrado: {excel_path}")
            print("📂 Conteúdo da pasta relatorios:")
            if os.path.exists("relatorios"):
                for item in os.listdir("relatorios"):
                    print(f"   - {item}")
            return
        
        print("✅ Arquivo Excel encontrado!")
        
        # Criar pasta do banco se não existir
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        
        # Testar verificação de estrutura
        print("\n1. Verificando estrutura do Excel...")
        resultado, mensagem = verificar_estrutura_excel(excel_path)
        print(f"📋 {mensagem}")
        
        if resultado:
            # Testar obtenção de colunas
            print("\n2. Obtendo colunas do Excel...")
            colunas = obter_colunas_excel(excel_path)
            print(f"📊 Colunas encontradas: {colunas}")
            
            # Testar importação
            print("\n3. Iniciando importação...")
            sucesso, mensagem = importar_excel_para_sqlite(excel_path, 'Estoque', db_path)
            print(f"🎯 Resultado: {'SUCESSO' if sucesso else 'FALHA'}")
            print(f"💬 {mensagem}")
        else:
            print("❌ Estrutura inválida, importação cancelada.")
            
    except Exception as e:
        print(f"💥 Erro no teste: {e}")
        import traceback
        traceback.print_exc()

# Executar teste se o arquivo for executado diretamente
if __name__ == "__main__":
    testar_importacao()
